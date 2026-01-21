#!/usr/bin/env python3
"""
Bond cross-check (v2):
- Reads inputs + Excel-cached outputs from tool/bond_toolkit.xlsx
- Recomputes the same metrics using the Python engine (flat-yield convention)
- Asserts tolerances (simple unit-test style)

Usage:
  python bond_check.py --xlsx ../tool/bond_toolkit.xlsx

Note:
- openpyxl reads cached formula values. If you edit the workbook, open it in Excel and Save once so cached values are updated.
"""
from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

import sys
sys.path.append(str(Path(__file__).resolve().parent))

from engine.bond_engine import BondSpec, FlatYield, build_flat_curve, cashflows, price_from_curve, price, macaulay_duration, modified_duration_flat, convexity_flat


@dataclass(frozen=True)
class ExcelOutputs:
    price: float
    mac_dur: float
    mod_dur: float
    convexity: float


def read_excel_outputs(xlsx_path: str, sheet: str = "Bond Toolkit") -> ExcelOutputs:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]
    return ExcelOutputs(
        price=float(ws["B13"].value),
        mac_dur=float(ws["B14"].value),
        mod_dur=float(ws["B15"].value),
        convexity=float(ws["B16"].value),
    )


def close(a: float, b: float, rel: float = 1e-9, abs_: float = 1e-6) -> bool:
    return abs(a - b) <= max(abs_, rel * max(abs(a), abs(b)))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--rel", type=float, default=1e-9)
    ap.add_argument("--abs", dest="abs_", type=float, default=1e-6)
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).resolve()

    wb = load_workbook(str(xlsx_path), data_only=True)
    ws = wb["Bond Toolkit"]
    bond = BondSpec(
        fv=float(ws["B5"].value),
        coupon_rate=float(ws["B6"].value),
        maturity=float(ws["B7"].value),
        coupon_freq=int(ws["B8"].value),
    )
    y = FlatYield(j=float(ws["B9"].value), m=int(ws["B10"].value))

    excel = read_excel_outputs(str(xlsx_path))

    curve = build_flat_curve(y)
    cf = cashflows(bond, stub=False)  # mirror original template (integer schedule)
    cf_pv = price_from_curve(cf, curve)
    P = price(cf_pv)
    mac = macaulay_duration(cf_pv)
    mod = modified_duration_flat(mac, y)
    conv = convexity_flat(cf_pv, y)

    print("=== Inputs ===")
    print(bond, y)
    print("\n=== Excel outputs (cached) ===")
    print(excel)
    print("\n=== Python recompute ===")
    print({"price": P, "mac_dur": mac, "mod_dur": mod, "convexity": conv})
    print("\n=== Differences (Python - Excel) ===")
    print({
        "price": P - excel.price,
        "mac_dur": mac - excel.mac_dur,
        "mod_dur": mod - excel.mod_dur,
        "convexity": conv - excel.convexity,
    })

    assert close(P, excel.price, rel=args.rel, abs_=args.abs_), "Price mismatch"
    assert close(mac, excel.mac_dur, rel=args.rel, abs_=1e-6), "Macaulay duration mismatch"
    assert close(mod, excel.mod_dur, rel=args.rel, abs_=1e-6), "Modified duration mismatch"
    assert close(conv, excel.convexity, rel=args.rel, abs_=1e-6), "Convexity mismatch"

    print("\nOK: Excel and Python match within tolerance.")


if __name__ == "__main__":
    main()
