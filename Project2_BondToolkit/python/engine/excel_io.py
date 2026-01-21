#!/usr/bin/env python3
"""
Excel IO helpers for Project 2 (v2).

- Reads the original template inputs from sheet "Bond Toolkit" cells:
    FV (B5), c (B6), T (B7), p (B8), j (B9), m (B10)

- Optionally reads a curve from a sheet called "Curve":
    Columns A:B starting at row 2:
      A: Tenor (years)
      B: Zero rate j^(m) (decimal)
    The compounding m is taken from "Bond Toolkit"!B10.

- Writes outputs to a new workbook, OR injects/overwrites a sheet named "Python Outputs"
  in an existing workbook.
"""
from __future__ import annotations

from dataclasses import asdict
from typing import Optional, Tuple, Dict
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .bond_engine import BondSpec, FlatYield, YieldCurve


DEFAULT_SHEET = "Bond Toolkit"


def read_inputs_xlsx(xlsx_path: str, sheet: str = DEFAULT_SHEET) -> Tuple[BondSpec, FlatYield]:
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {xlsx_path}")
    ws = wb[sheet]

    fv = float(ws["B5"].value)
    c = float(ws["B6"].value)
    T = float(ws["B7"].value)
    p = int(ws["B8"].value)
    j = float(ws["B9"].value)
    m = int(ws["B10"].value)

    return BondSpec(fv=fv, coupon_rate=c, maturity=T, coupon_freq=p), FlatYield(j=j, m=m)


def read_curve_xlsx(xlsx_path: str, curve_sheet: str = "Curve", sheet: str = DEFAULT_SHEET) -> Optional[YieldCurve]:
    wb = load_workbook(xlsx_path, data_only=True)
    if curve_sheet not in wb.sheetnames:
        return None
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found in {xlsx_path}")
    ws_curve = wb[curve_sheet]
    ws_base = wb[sheet]
    m = int(ws_base["B10"].value)

    pillars = []
    rates = []
    # expect header in row 1, start at row 2
    for r in range(2, 5000):
        t = ws_curve[f"A{r}"].value
        y = ws_curve[f"B{r}"].value
        if t is None and y is None:
            break
        if t is None or y is None:
            # skip partially empty rows
            continue
        pillars.append(float(t))
        rates.append(float(y))

    if not pillars:
        return None
    return YieldCurve(tuple(pillars), tuple(rates), m)


def write_output_workbook(
    out_xlsx: str,
    summary: Dict[str, float],
    cashflows: pd.DataFrame,
    scenarios: Optional[pd.DataFrame] = None,
    key_rate: Optional[pd.DataFrame] = None,
) -> None:
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        pd.DataFrame([summary]).to_excel(writer, sheet_name="Summary", index=False)
        cashflows.to_excel(writer, sheet_name="Cashflows", index=False)
        if scenarios is not None:
            scenarios.to_excel(writer, sheet_name="Scenarios", index=False)
        if key_rate is not None:
            key_rate.to_excel(writer, sheet_name="KeyRateDur", index=False)


def inject_python_outputs(
    xlsx_in: str,
    xlsx_out: str,
    summary: Dict[str, float],
    cashflows: pd.DataFrame,
    scenarios: Optional[pd.DataFrame] = None,
    key_rate: Optional[pd.DataFrame] = None,
    sheet_name: str = "Python Outputs",
) -> None:
    """
    Copy an existing workbook and overwrite/create a sheet for Python outputs.
    """
    wb = load_workbook(xlsx_in)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)

    # write summary starting at A1
    ws["A1"] = "Python Engine Outputs"
    ws["A3"] = "Summary"
    # header
    headers = list(summary.keys())
    for j, h in enumerate(headers, start=1):
        ws.cell(row=4, column=j, value=h)
        ws.cell(row=5, column=j, value=float(summary[h]))

    # cashflows
    start_row = 7
    ws.cell(row=start_row, column=1, value="Cashflows")
    df = cashflows.copy()
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row+1, column=j, value=col)
    for i, row in enumerate(df.itertuples(index=False), start=0):
        for j, val in enumerate(row, start=1):
            ws.cell(row=start_row+2+i, column=j, value=float(val) if isinstance(val, (int,float)) else val)

    cur_row = start_row + 2 + len(df) + 2

    def write_df(title: str, df2: Optional[pd.DataFrame]):
        nonlocal cur_row
        if df2 is None:
            return
        ws.cell(row=cur_row, column=1, value=title)
        for j, col in enumerate(df2.columns, start=1):
            ws.cell(row=cur_row+1, column=j, value=col)
        for i, row in enumerate(df2.itertuples(index=False), start=0):
            for j, val in enumerate(row, start=1):
                ws.cell(row=cur_row+2+i, column=j, value=float(val) if isinstance(val,(int,float)) else val)
        cur_row += 2 + len(df2) + 2

    write_df("Scenarios", scenarios)
    write_df("KeyRateDurations", key_rate)

    wb.save(xlsx_out)
